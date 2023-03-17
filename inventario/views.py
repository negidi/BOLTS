from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from .models import activo
from .models import herramientas
from .models import maquinaria
from .models import Usuario
from .models import clientes
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, permission_required ##roles##
from django.contrib.auth.models import Group ##
from django.contrib import messages
from datetime import datetime
from datetime import date
from .admin import clientesAdmin
from openpyxl import Workbook
from django.http.response import HttpResponse
from auditlog.models import LogEntry
import json
import re 




from inventario.forms import FormularioRegistrar
# Create your views here.




def login_request(request) :
    if request.user.is_authenticated: 
         ##si el usuario esta autenticado redirige##
        return redirect("/home")
    
        
    mensaje = ""
    form = AuthenticationForm()
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            usuario = form.cleaned_data.get("username")
            contrasenna = form.cleaned_data.get("password")
            user = authenticate(username=usuario, password=contrasenna)
            if user is not None:
                userAuth = Usuario.objects.get(username=usuario)
                if userAuth.username == 'admin':
                    userAuth.rol = 'SUPER_ADMINISTRADOR'
                request.session['ROL'] = userAuth.rol
                
                login(request, user)  # para loguear al usuario#
                messages.info(request, f"{usuario} logueado")
                return redirect("/")
            else:
                mensaje = "Error de usuario o contraseña"
                messages.error(request, "Error de usuario o contraseña")
        else:
            mensaje = "Error de usuario o contraseña"
            messages.error(request, "Error de usuario o contraseña")
    return render(request, "registration/login.html", {"mensaje": mensaje})

@login_required ##permisosderoles##
@permission_required("inventario.usuario",raise_exception=True)
def listarusuario(request):
    x = Usuario.objects.all()
    return render(request,"listarusuario.html", {"usuario" : x})

@login_required ##permisosderoles##
@permission_required("inventario.session",raise_exception=True)
def signup(request):
    mensaje = ""
    if request.method == "POST":
        form = FormularioRegistrar(request.POST)
        print(form)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get("username")
            raw_password = form.cleaned_data.get("password1")
            user = authenticate(username=username, password=raw_password)
            login(request, user)
            mensaje = "Usuario creado correctamente"
        else:
            mensaje = "Favor validar contraseña, considerar el largo de 8 caracteres y no ingresar contraseñas anteriores"
    else:
        form = FormularioRegistrar()

    return render(request, "registration/signup.html", {"form": form, "mensaje": mensaje})

@permission_required("inventario.session",raise_exception=True)
def eliminarusuarios(request):
    return render(request, "eliminarusuarios.html")

@login_required ## Auditoria##
@permission_required("inventario.delete_user",raise_exception=True)
def eliminarus(request):
    us = None
    mensaje = ""

    try:
        us = Usuario.objects.get(username= request.GET["username"])
        return render(request, "eliminarusuarios.html", {"us":us})

    except: 
        us = None
    
    if us == None:
        username = None
        try:
            username = request.POST["username"]
        except:
            username = None

        if username != None:
            us = Usuario.objects.get(username = username)

            nombres_a = request.POST["nombre_usuario"]
            apellidos_a = request.POST["primer_apellido"]
            rut_a = request.POST["rut_usuario"]
            email_a = request.POST["email_usuario"]
            rol_a = request.POST["rol"]

            us.first_name = nombres_a
            us.last_name= apellidos_a
            us.rut= rut_a
            us.email = email_a
            ##telefono_usuario = request.POST["telefono_usuario "]##
            us.rol = rol_a
            us.is_active = False
            try:
                us.save()
                mensaje = "Se ha Eliminado el usuario"
            except:
                mensaje = "Ha ocurrido un error al eliminar"


            return render(request, "eliminarusuarios.html", {"mensaje":mensaje})
        
        else:
            mensaje = "No se encontro el usuario ingresado"

            return render(request, "eliminarusuarios.html", {"mensaje":mensaje})
    else:
        mensaje = "No se encontró usuario"

        return render(request, "eliminarusuarios.html", {"mensaje":mensaje})

def logout_request(request):
    logout(request)
    return redirect("/auth/login_request")

# Create your views here.

def index(request):
    return render(request, 'index.html')

@login_required ##
def home(request):
    return render(request, 'home.html')


@login_required ##
@permission_required("inventario.activo",raise_exception=True)
def creacionActivo(request):

    return render(request, 'creacionActivo.html')

@login_required ##
@permission_required("inventario.activo",raise_exception=True)
def crearHerramientas(request):
    id_activo_automatic = str((herramientas.objects.all().count() + 1)  * 7)
    idid = "herr" +  id_activo_automatic
    return render(request, 'crearHerramientas.html',{'idid':idid})

@login_required ##
@permission_required("inventario.add_herramientas",raise_exception=True)
def registrarHerramienta(request):
    precio= None
    mensaje=""

    id_activo_automatic = str((herramientas.objects.all().count() + 1)  * 7)
    idid = "herr" +  id_activo_automatic
    mensaje = "Herramienta ingresada correctamente"
    nombre_herramientas = request.POST['nombre_activo']
    cantidad_herramientas = request.POST['cantidad']

    if cantidad_herramientas != 0:
        estado_herramienta = 'Con Stock'
        cantidad_herramientas = request.POST['cantidad']
    else:
        estado_herramienta = 'Sin Stock'
        mensaje = "el precio debe ser mayor a 0"  

    precio_herramientas = request.POST['precio']
    if precio_herramientas != 0:
        precio_herramientas = request.POST['precio']        
    else:
        mensaje = "el precio debe ser mayor a 0"  
     
    proveedor_herramientas = request.POST['proveedor']
    marca_herramientas = request.POST ['marca']

    x = activo.objects.create(id_activo = idid , nombre_activo = nombre_herramientas , cantidad = cantidad_herramientas , precio = precio_herramientas, estado = estado_herramienta, activo = 'Activo')
    herramientas.objects.create(proveedor = proveedor_herramientas, marca = marca_herramientas, activo = x)    

    return render(request, 'respuestaHtas.html',{"mensaje": mensaje})

@permission_required("inventario.session",raise_exception=True)
def eliminarherramientas(request):
    return render(request,"eliminarherramientas.html")

@login_required ## Auditoria##
@permission_required("inventario.delete_herramientas}",raise_exception=True)
def eliminarherr(request):

    herr = None
    mensaje = ""
    try:
        herr = herramientas.objects.get(activo = request.GET["idactivo"])
        act = activo.objects.get(id_activo = request.GET["idactivo"])
        return render(request, "eliminarherramientas.html", {"herra":herr, "acti":act})
    except:
        herr = None
    
    if herr == None:
        id_activo = None
        try:
            id_activo = request.POST["id_activo"]
        except:
            id_activo = None

        if id_activo != None:
            herr = herramientas.objects.get(activo = id_activo)
            act = activo.objects.get(id_activo = id_activo)

            nombre_a = request.POST["nombre_maquina"]
            cantidad_a = request.POST["cantidad"]
            precio_a = request.POST["precio"]
            estado_a = request.POST["estado"]
            proveedorherr_a = request.POST["proveedorherr"]
            marcaherr_a = request.POST["marcaherr"]
            act.nombre_activo = nombre_a
            act.cantidad = cantidad_a
            act.precio = precio_a
            act.estado = estado_a
            act.activo = 'Inactivo'
            herr.proveedor = proveedorherr_a
            herr.marca = marcaherr_a
            
            try:
                act.save()
                herr.save()
                mensaje = "Herramienta eliminada con exito"
            except:
                mensaje = "Ha ocurrido un error al Elimnar la herramienta"


            return render(request, "eliminarherramientas.html", {"mensaje":mensaje})
        
        else:
            mensaje = "No se ha encontrado la herramienta ingresada"

            return render(request, "eliminarherramientas.html", {"mensaje":mensaje})
    else:
        mensaje = "No se encontró la herramienta ingresada"

        return render(request, "eliminarherramientas.html", {"mensaje":mensaje})




@login_required ##
@permission_required("inventario.activo",raise_exception=True)
def crearMaquinaria(request):

    id_activo_automatic = str((maquinaria.objects.all().count() + 1)  * 7)
    idid = "maqu" +  id_activo_automatic
    return render(request, 'crearMaquinaria.html',{'idid':idid})


@login_required ##
@permission_required("inventario.add_maquinaria",raise_exception=True)##
def registrarMaquinaria(request):

    mensaje=""
    id_activo_automatic = str((maquinaria.objects.all().count() + 1)  * 7)
    idid = "maqu" +  id_activo_automatic
    mensaje = "Maquinaria ingresada correctamente"
    nombre_mqnas = request.POST['nombre_activo']
    cantidad_mqnas = request.POST['cantidad']
    if cantidad_mqnas != 0:
        estado_mqnas = 'Con Stock'
        cantidad_mqnas = request.POST['cantidad']
    else:
        estado_mqnas = 'Sin Stock'
        mensaje = "el precio debe ser mayor a 0" 

    precio_mqnas = request.POST['precio']
    if precio_mqnas != 0:
        precio_mqnas = request.POST['precio']        
    else:
        mensaje = "el precio debe ser mayor a 0"  

    diseno_mqnas = request.FILES['disenomaq']    
    estruc_mqnas = request.POST ['estructuramaq']


    y = activo.objects.create(id_activo = idid , nombre_activo = nombre_mqnas , cantidad = cantidad_mqnas , precio = precio_mqnas, estado = estado_mqnas, activo = 'Activo') 
    maquinaria.objects.create(diseño_maquinaria = diseno_mqnas, material_estructura = estruc_mqnas, activo = y)

    return render(request, 'respuestahtas.html',{"mensaje": mensaje})

@permission_required("inventario.session",raise_exception=True)
def eliminarmaquinarias(request):
    return render(request, "eliminarmaquinarias.html")

@login_required ## Auditoria##
@permission_required("inventario.delete_maquinaria",raise_exception=True)
def eliminarmaq(request):

    maq = None
    mensaje = " "
    try:
        maq = maquinaria.objects.get(activo = request.GET["idactivo"])
        act = activo.objects.get(id_activo = request.GET["idactivo"])
        return render(request, "eliminarmaquinarias.html", {"maqui":maq, "acti":act})
    except:
        maq =None 

    if maq == None:
        id_activo = None
        try:
            id_activo = request.POST["id_activo"]
        except:
            id_activo = None

        if id_activo != None:
            maq = maquinaria.objects.get(activo = id_activo)
            act = activo.objects.get(id_activo = id_activo)

            nombre_a = request.POST["nombre_maquina"]
            cantidad_a = request.POST["cantidad"]
            precio_a = request.POST["precio"]
            disenomaq_a = request.POST["disenomaq"]
            estructuramaq_a = request.POST["estructuramaq"]
            act.nombre_activo = nombre_a
            act.cantidad = cantidad_a
            act.precio = precio_a
            act.estado = 'Inacativo'
            maq.diseño_maquinaria = disenomaq_a
            maq.material_estructura = estructuramaq_a
            
            try:
                act.save()
                maq.save()
                mensaje = "Maquina eliminada con exito"
            except:
                mensaje = "Ha ocurrido un error al eliminar la maquinaria"

            return render(request, "eliminarmaquinarias.html", {"mensaje":mensaje})
        
        else:
            mensaje = "No se ha encontrado la maquina ingresada"

            return render(request, "eliminarmaquinarias.html", {"mensaje":mensaje})
    else:
        mensaje = "No se encontró la maquina ingresada"

        return render(request, "eliminarmaquinarias.html", {"mensaje":mensaje})   

def actualizarmaq(request):
    return render(request,"actualizarmaquina.html")

def actualizarherr(request):
    return render(request,"actualizarherramientas.html")

@login_required ##
@permission_required("inventario.change_maquinaria",raise_exception=True)
def editarmaq(request):
    maq = None
    mensaje = ""
    try:
        maq = maquinaria.objects.get(activo = request.GET["idactivo"])
        act = activo.objects.get(id_activo = request.GET["idactivo"])
        if act.activo == "Activo":
            return render(request, "actualizarmaquina.html", {"maqui":maq, "acti":act})
        else:
            mensaje = 'No se encontro la maquinaria ingresada'
            return render(request, "actualizarmaquina.html", {"mensaje":mensaje} )
    except:
        maq = None
    
    if maq == None:
        id_activo = None
        try:
            id_activo = request.POST["id_activo"]
        except:
            id_activo = None

        if id_activo != None and id_activo != "":
            maq = maquinaria.objects.get(activo = id_activo)
            act = activo.objects.get(id_activo = id_activo)

            nombre_a = request.POST["nombre_maquina"]
            cantidad_a = request.POST["cantidad"]
            precio_a = request.POST["precio"]
            disenomaq_a = request.FILES["disenomaq"]
            estructuramaq_a = request.POST["estructuramaq"]
            if cantidad_a != '0':
                estado_mqnas = 'Con Stock'
            else:
                estado_mqnas = 'Sin Stock'

            act.nombre_activo = nombre_a
            act.cantidad = cantidad_a
            act.precio = precio_a
            act.estado = estado_mqnas
            maq.diseño_maquinaria = disenomaq_a
            maq.material_estructura = estructuramaq_a
            
            try:
                act.save()
                maq.save()
                mensaje = "Maquina actualizada con exito"
            except:
                mensaje = "Ha ocurrido un error al actualizar la maquinaria"

            return render(request, "actualizarmaquina.html", {"mensaje":mensaje})
        
        else:
            mensaje = "No se ha encontrado la maquina ingresada"

            return render(request, "actualizarmaquina.html", {"mensaje":mensaje})
    else:
        mensaje = "No se encontró la maquina ingresada"

        return render(request, "actualizarmaquina.html", {"mensaje":mensaje})   

@login_required ##
@permission_required("inventario.change_herramientas",raise_exception=True)
def editarherr(request):
    herr = None
    mensaje = ""
    try:
        herr = herramientas.objects.get(activo = request.GET["idactivo"])
        act = activo.objects.get(id_activo = request.GET["idactivo"])
        if act.activo == "Activo":
            return render(request, "actualizarherramientas.html", {"herra":herr, "acti":act})
        else:
            mensaje = 'No se encontro la herramienta ingresada'
            return render(request, "actualizarherramientas.html", {"mensaje":mensaje})
    except:
        herr = None
    
    if herr == None:
        id_activo = None
        try:
            id_activo = request.POST["id_activo"]
        except:
            id_activo = None

        if id_activo != None and id_activo != "":
            herr = herramientas.objects.get(activo = id_activo)
            act = activo.objects.get(id_activo = id_activo)

            nombre_a = request.POST["nombre_maquina"]
            cantidad_a = request.POST["cantidad"]
            precio_a = request.POST["precio"]
            proveedorherr_a = request.POST["proveedorherr"]
            marcaherr_a = request.POST["marcaherr"]
            if cantidad_a != '0':
                estado_herramienta = 'Con Stock'
            else:
                estado_herramienta = 'Sin Stock'

            act.nombre_activo = nombre_a
            act.cantidad = cantidad_a
            act.precio = precio_a
            act.estado = estado_herramienta
            herr.proveedor = proveedorherr_a
            herr.marca = marcaherr_a        
            

            try:
                act.save()
                herr.save()
                mensaje = "Herramienta actualizada con exito"
            except:
                mensaje = "Ha ocurrido un error al actualizar la herramienta"


            return render(request, "actualizarherramientas.html", {"mensaje":mensaje})
        
        else:
            mensaje = "No se ha encontrado la herramienta ingresada"

            return render(request, "actualizarherramientas.html", {"mensaje":mensaje})
    else:
        mensaje = "No se encontró la herramienta ingresada"

        return render(request, "actualizarherramientas.html", {"mensaje":mensaje})
          
def respuestaherramientas(request):
    return render(request, "respuestaherramientas.html")

@login_required ##
@permission_required("inventario.activo",raise_exception=True)
def crearclientes(request):
    id_activo_automatic = str((clientes.objects.all().count() + 1)  * 7)
    idco = "idco" +  id_activo_automatic
    y = maquinaria.objects.all()
    return render(request,"crearClientes.html",{'y': y, 'idco':idco})


@login_required ##
@permission_required("inventario.add_clientes",raise_exception=True)
def registrarCliente(request):

    mensaje = ""
    id_activo_automatic = str((clientes.objects.all().count() + 1)  * 7)
    idco = "idco" +  id_activo_automatic
    mensaje = "Cliente ingresado correctamente"
    rut_clte = request.POST['rut_cliente']
    nombre_clte= request.POST['nombre_cliente']
    telefono_clt = request.POST['telefono_contacto']
    direccion_clte = request.POST['direccion']
    
    unidad = request.POST['unidades']

    

    if unidad != 0:
        unidad = request.POST['unidades']
    else:
        mensaje = "la unidad debe ser mayor a 0"  

    if rut_clte[0:8].isnumeric():
        print("Paso a nivel 0")
        if len(rut_clte) == 10:
            print("Paso a nivel 1")
            if rut_clte.find("-") == 8:
                print("Paso a nivel 2")
 
            else:
                mensaje = "Ingreso mal el rut"
                return render(request, 'respuestaHtas.html',{"mensaje": mensaje})
        else:
            mensaje = "largo de rut no corresponde"
            return render(request, 'respuestaHtas.html',{"mensaje": mensaje})
    else:
            mensaje = "Ingreso letras"
            return render(request, 'respuestaHtas.html',{"mensaje": mensaje})
    
    if telefono_clt.isnumeric():
        pass
    else:
        mensaje = "El telefono solo debe contener numeros"
        return render(request, 'respuestaHtas.html',{"mensaje": mensaje})

    fecha = request.POST['fecha']
    maqui = request.POST['cod_maquina']
    z = maquinaria.objects.get(activo__maquinaria__activo = maqui)
    total = int(unidad) * int(z.activo.precio)
    clientes.objects.create(id_compra = idco,rut_cliente = rut_clte, nombre_cliente = nombre_clte, telefono_contacto =telefono_clt, direccion = direccion_clte, unidades = unidad, fecha = fecha ,maquinaria = z,total = total,activo = 'Activo')

    return render(request, 'respuestaHtas.html',{"mensaje": mensaje})


@login_required ##
def listarherramientas(request):

    x = herramientas.objects.select_related ('activo').all()
    p = herramientas.objects.all()
    return render(request,"listarherramientas.html", {'herramientas':x, "p" : p})


@login_required ##
def listarmaquinaria(request):

    y = maquinaria.objects.select_related ('activo').all()

    return render(request,"listarmaquinaria.html", {'maquinaria': y})


@login_required ##
def listarclientes(request):

    z = clientes.objects.select_related ('maquinaria').all()
    

    return render(request,"listarclientes.html", {'clientes': z})

def filtroherr(request):
    if request.GET["nombre_activ"] and request.GET["prec"] and request.GET["provee"]: # valida por nombre y precio y proveedor (variacion en los precios.)
        prov = request.GET["provee"]
        htas = request.GET["nombre_activ"]
        if request.GET["validar"] == 'pmayor':
            prec = request.GET["prec"]
            y = herramientas.objects.filter(activo__precio__gt=prec, activo__nombre_activo__iexact=htas, proveedor__iexact=prov)
        elif request.GET["validar"] == 'pmenor':
            prec = request.GET["prec"]
            y = herramientas.objects.filter(activo__precio__lt=prec, activo__nombre_activo__iexact=htas, proveedor__iexact=prov)
        elif request.GET["validar"] == 'pigual':
            prec = request.GET["prec"]
            y = herramientas.objects.filter(activo__precio__exact=prec, activo__nombre_activo__iexact=htas, proveedor__iexact=prov)
        p = herramientas.objects.all()
        return render(request,"listarherramientas.html",{"fullhtas": y,"query":prec,'p':p})
    
    elif request.GET["nombre_activ"] and request.GET["prec"]: # valida por nombre y precio(variable)
        htas = request.GET["nombre_activ"]
        if request.GET["validar"] == 'pmayor':
                prec = request.GET["prec"]
                y = herramientas.objects.filter(activo__precio__gt=prec,activo__nombre_activo__iexact=htas)
        elif request.GET["validar"] == 'pmenor':
            prec = request.GET["prec"]
            y = herramientas.objects.filter(activo__precio__lt=prec,activo__nombre_activo__iexact=htas)
        elif request.GET["validar"] == 'pigual':
            prec = request.GET["prec"]
            y = herramientas.objects.filter(activo__precio__exact=prec,activo__nombre_activo__iexact=htas)
        p = herramientas.objects.all()
        return render(request,"listarherramientas.html",{"fullhtas": y,"query":prec,'p':p})
    
    elif request.GET["nombre_activ"] and request.GET["provee"]: # valida por nombre y proveedor
        htas = request.GET["nombre_activ"]
        prov = request.GET["provee"]
        p = herramientas.objects.all()
        y = herramientas.objects.filter(activo__nombre_activo__iexact = htas , proveedor__iexact=prov)
        return render(request,"listarherramientas.html",{"fullhtas": y,'p':p})
    
    elif request.GET["prec"] and request.GET["provee"]: # valida por precio y proveedor
        prov = request.GET["provee"]
        prec = request.GET["prec"]
        p = herramientas.objects.all()
        y = herramientas.objects.filter(activo__precio__exact=prec, proveedor__iexact=prov)
        return render(request,"listarherramientas.html",{"fullhtas": y,"query":prec,'p':p})
    
    elif request.GET["nombre_activ"]: # valida el nombre del activo
        htas = request.GET["nombre_activ"]
        p = herramientas.objects.all()
        x = herramientas.objects.filter(activo__nombre_activo__icontains=htas)
        
        return render(request,"listarherramientas.html",{"fullhtas": x,"query":htas,'p':p})
    elif request.GET["provee"] != '': # valida el nombre del activo
        htas = request.GET["provee"]
        x = herramientas.objects.filter(proveedor__iexact=htas)
        p = herramientas.objects.all()
        return render(request,"listarherramientas.html",{"fullhtas": x,"query":htas, 'p':p})
  
    elif request.GET["prec"]: #Valida el precio mayor, menor o igual.
        if request.GET["validar"] == 'pmayor':
                prec = request.GET["prec"]
                x = herramientas.objects.filter(activo__precio__gt=prec)
                return render(request,"listarherramientas.html",{"fullhtas": x,"query":prec})
        elif request.GET["validar"] == 'pmenor':
            prec = request.GET["prec"]
            x = herramientas.objects.filter(activo__precio__lt=prec)
            return render(request,"listarherramientas.html",{"fullhtas": x,"query":prec})
        elif request.GET["validar"] == 'pigual':
            prec = request.GET["prec"]
            x = herramientas.objects.filter(activo__precio__exact=prec)
            return render(request,"listarherramientas.html",{"fullhtas": x,"query":prec}) 
    else:
        mensaje = "Debe ingresar un nombre de herramienta"
        return render(request,"listarherramientas.html",{"mensaje": mensaje})
        
def filtromaqu(request):
    if request.GET["nombre_activ"] and request.GET["prec"]: # valida por nombre y precio(exacto)
        htas = request.GET["nombre_activ"]
        if request.GET["validar"] == 'pmayor':
            prec = request.GET["prec"]
            x = maquinaria.objects.filter(activo__precio__gt=prec, activo__nombre_activo__iexact=htas)
        elif request.GET["validar"] == 'pmenor':
            prec = request.GET["prec"]
            x = maquinaria.objects.filter(activo__precio__lt=prec, activo__nombre_activo__iexact=htas)
        elif request.GET["validar"] == 'pigual':
            prec = request.GET["prec"]
            x = maquinaria.objects.filter(activo__precio__exact=prec, activo__nombre_activo__iexact=htas)
        return render(request,"listarmaquinaria.html",{"fullhtas": x,"query":prec})
    elif request.GET["nombre_activ"]: # valida el nombre del activo
        htas = request.GET["nombre_activ"]
        x = maquinaria.objects.filter(activo__nombre_activo__iexact=htas)
        return render(request,"listarmaquinaria.html",{"fullhtas": x,"query":htas})
    elif request.GET["prec"]: #Valida el precio mayor, menor o igual.
        if request.GET["validar"] == 'pmayor':
            prec = request.GET["prec"]
            x = maquinaria.objects.filter(activo__precio__gt=prec)
            return render(request,"listarmaquinaria.html",{"fullhtas": x,"query":prec})
        elif request.GET["validar"] == 'pmenor':
            prec = request.GET["prec"]
            x = maquinaria.objects.filter(activo__precio__lt=prec)
            return render(request,"listarmaquinaria.html",{"fullhtas": x,"query":prec})
        elif request.GET["validar"] == 'pigual':
            prec = request.GET["prec"]
            x = maquinaria.objects.filter(activo__precio__exact=prec)
            return render(request,"listarmaquinaria.html",{"fullhtas": x,"query":prec}) 
    

    else:
        mensaje = "Debe ingresar un nombre de herramienta"
        return render(request,"listarmaquinaria.html",{"mensaje": mensaje})
        
def filtroclnte(request):
    if request.GET["nombre_cliente"] and  request.GET["fecha"]: #Validando nombre de cliente y fecha de registro
        fullclientes = request.GET["nombre_cliente"] # Get nombre
        f = request.GET["fecha"]# Get fecha
        x = clientes.objects.filter(nombre_cliente__iexact= fullclientes, fecha = f) #Busqueda en BD
        w=x.count() #Contador de resgistros para el cliente buscado

        return render(request,"listarclientes.html",{"fullclte": x,"query":fullclientes, "contador": w})
    elif request.GET["nombre_cliente"]:
        fullclientes = request.GET["nombre_cliente"] # Get nombre
        x = clientes.objects.filter(nombre_cliente__iexact = fullclientes) #comparacion del nombre ingresado
        w=x.count() #Contador de resgistros para el cliente buscado
        return render(request, "listarclientes.html", {"fullclte" : x,"query":fullclientes, "contador": w})
    elif request.GET["fecha"]: #validar comparacion de fechas 
        if request.GET["validar"] == 'figual':
            f = request.GET["fecha"]# Fecha igual
            x = clientes.objects.filter(fecha__exact = f) 
            w=x.count() #Contador de resgistros para el cliente buscado
            return render(request, "listarclientes.html", {"fullclte" : x, "contador": w})
        elif request.GET["validar"] == 'fmayor':
            f = request.GET["fecha"]# Fecha mayor a 
            x = clientes.objects.filter(fecha__gt=f) 
            w=x.count() #Contador de resgistros para el cliente buscado
            return render(request, "listarclientes.html", {"fullclte" : x, "contador": w})
        else:
            f = request.GET["fecha"]# Fecha menor a 
            x = clientes.objects.filter(fecha__lt=f) 
            w=x.count() #Contador de resgistros para el cliente buscado
            return render(request, "listarclientes.html", {"fullclte" : x, "contador": w})

    #Crear filtro por las maquinas
    #elif request.GET["cod_maquina"]:
    #    x = request.GET["cod_maquina"]
    #    y = maquinaria.objects.all()
    #    return render(request,"listarclientes.html",{"y": y, })
        
    else:
        mensaje = "Debe ingresar un nombre de Cliente"
        return render(request,"listarclientes.html",{"mensaje": mensaje})

@login_required ##
def reporteclientes(request):
        
        z = clientes.objects.select_related ('maquinaria').all()
        
        wb = Workbook()
        ws = wb.active
        ws["B1"] = "Reporte de Clientes"

        ws.merge_cells("B1:E1")
        ws["B3"]  = "id_compra"
        ws["C3"]  = "rut_cliente"
        ws["D3"]  = "nombre_cliente"
        ws["E3"]  = "telefono_contacto"
        ws["F3"]  = "direccion"
        ws["G3"]  = "unidades"
        ws["H3"]  = "fecha"
        ws["I3"]  = "total"
        ws["J3"]  = "activo"
        ws["K3"]  = "Nombre maquina"
        ws["L3"]  = "maquinaria"
        ws["M3"]  = "Precio"
        contador = 4
        for cliente in z:
            
            ws.cell(row = contador, column = 2).value = cliente.id_compra
            ws.cell(row = contador, column = 3).value = cliente.rut_cliente
            ws.cell(row = contador, column = 4).value = cliente.nombre_cliente
            ws.cell(row = contador, column = 5).value = cliente.telefono_contacto
            ws.cell(row = contador, column = 6).value = cliente.direccion
            ws.cell(row = contador, column = 7).value = cliente.unidades
            ws.cell(row = contador, column = 8).value = cliente.fecha
            ws.cell(row = contador, column = 9).value = cliente.total
            ws.cell(row = contador, column = 10).value = cliente.activo
            ws.cell(row = contador, column = 11).value = cliente.maquinaria.activo. nombre_activo
            ws.cell(row = contador, column = 12).value = cliente.maquinaria.material_estructura
            ws.cell(row = contador, column = 13).value = cliente.maquinaria.activo.precio
            contador+=1
        nombre_archivo = "reporteClientes.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return(response)


@login_required ##
def reporteMaquinaria(request):
        
        y = maquinaria.objects.select_related ('activo').all()
        wb = Workbook()
        ws = wb.active
        ws["B1"] = "Reporte de Maquinas"

        ws.merge_cells("B1:E1")
        ws["B3"]  = "ID Activo"
        ws["C3"]  = "Nombre"
        ws["D3"]  = "Cantidad"
        ws["E3"]  = "Precio"
        ws["F3"]  = "Material Estructura"
        ws["G3"]  = "Estado"
        ws["H3"]  = "Stock"

        contador = 4
        for Maquina in y:
            
            ws.cell(row = contador, column = 2).value = Maquina.activo.id_activo
            ws.cell(row = contador, column = 3).value = Maquina.activo.nombre_activo
            ws.cell(row = contador, column = 4).value = Maquina.activo.cantidad
            ws.cell(row = contador, column = 5).value = Maquina.activo.precio
            ws.cell(row = contador, column = 6).value = Maquina.material_estructura
            ws.cell(row = contador, column = 7).value = Maquina.activo.activo
            ws.cell(row = contador, column = 8).value = Maquina.activo.estado
            contador+=1
        nombre_archivo = "reporteMaquinaria.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return(response)

@login_required ##
def reporteHerramienta(request):
        
        z = herramientas.objects.select_related ('activo').all()
        wb = Workbook()
        ws = wb.active
        ws["B1"] = "Reporte de Herramientas"

        ws.merge_cells("B1:E1")
        ws["B3"]  = "ID Activo"
        ws["C3"]  = "Nombre"
        ws["D3"]  = "Cantidad"
        ws["E3"]  = "Precio"
        ws["F3"]  = "Proveedor"
        ws["G3"]  = "Marca"
        ws["H3"]  = "Estado"
        ws["I3"]  = "Stock"

        contador = 4
        for Herramienta in z:
            
            ws.cell(row = contador, column = 2).value = Herramienta.activo.id_activo
            ws.cell(row = contador, column = 3).value = Herramienta.activo.nombre_activo
            ws.cell(row = contador, column = 4).value = Herramienta.activo.cantidad
            ws.cell(row = contador, column = 5).value = Herramienta.activo.precio
            ws.cell(row = contador, column = 6).value = Herramienta.proveedor
            ws.cell(row = contador, column = 7).value = Herramienta.marca
            ws.cell(row = contador, column = 8).value = Herramienta.activo.activo
            ws.cell(row = contador, column = 9).value = Herramienta.activo.estado
            contador+=1
        nombre_archivo = "reporteHerramientas.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return(response)

@login_required ##
def reporteUsuario(request):
        
        x = Usuario.objects.all()
        wb = Workbook()
        ws = wb.active
        ws["B1"] = "Reporte de Usuarios"

        ws.merge_cells("B1:E1")
        ws["B3"]  = "RUT"
        ws["C3"]  = "Nombre"
        ws["D3"]  = "Apellido"
        ws["E3"]  = "Email"
        ws["F3"]  = "Estado"
        ws["G3"]  = "ROL"

        contador = 4
        for usuario in x:
            
            ws.cell(row = contador, column = 2).value = usuario.rut
            ws.cell(row = contador, column = 3).value = usuario.first_name
            ws.cell(row = contador, column = 4).value = usuario.last_name
            ws.cell(row = contador, column = 5).value = usuario.email
            ws.cell(row = contador, column = 6).value = usuario.is_active
            ws.cell(row = contador, column = 7).value = usuario.rol
            contador+=1
        nombre_archivo = "reporteUsuario.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return(response)




@login_required ##
@permission_required("inventario.activo",raise_exception=True)
def actualizarcliente(request):
    return render(request, 'actualizarcliente.html')

@login_required ##
@permission_required("inventario.change_clientes",raise_exception=True)
def editarclte(request): #validar 
    clte = None
    mensaje = ""

    try:
        y = maquinaria.objects.all()
        
        clte = clientes.objects.get(id_compra = request.GET["idcliente"])
        if clte.activo == "Activo":
            return render(request, "actualizarcliente.html", {"cltes" : clte, "y": y})
        else:
            mensaje = 'No se encontro orden de compra ingresada'
            return render(request, "actualizarcliente.html", {"mensaje":mensaje} )
    except:
        clte = None


    if clte == None:
        id_compra = None
        try:
            id_compra = request.POST["id_compra"]
        except:
            id_compra = None

        if id_compra != None and id_compra != "":
            clte = clientes.objects.get(id_compra = id_compra)

            rut_clt = request.POST["rut_cliente"]
            nombre_clt = request.POST["nombre_cliente"]
            telefono_clt = request.POST["telefono_contacto"]
            direccion_clt = request.POST["direccion"] 
            unidades_clt = request.POST["unidades"]
            fecha_clt = datetime.now()
            total_clt = request.POST ["total"]
            maqui = request.POST['cod_maquina']
            z = maquinaria.objects.get(activo__maquinaria__activo = maqui)
            y = maquinaria.objects.all()
            
            clte.rut_cliente = rut_clt
            clte.nombre_cliente = nombre_clt
            clte.telefono_contacto = telefono_clt
            clte.direccion = direccion_clt
            clte.unidades = unidades_clt
            clte.fecha = fecha_clt
            clte.total = total_clt
            clte.maquinaria = z

            try:
                clte.save()
                mensaje = "Se ha actualizado el cliente"
            except:
                mensaje = "Ha ocurrido un error al actualizar el cliente"
            
            return render(request, "actualizarcliente.html", {"mensaje":mensaje,"y" : y })
        else:
            mensaje = "No se ha encontrado el registro consultado"

            return render(request, "actualizarcliente.html", {"mensaje":mensaje, "y": y})
    else:
        mensaje = "No se encontró el registro consultado"

        return render(request, "actualizarcliente.html", {"mensaje":mensaje})

@permission_required("inventario.session",raise_exception=True)
def eliminarclientes(request):
    return render(request, "eliminarclientes.html")

@login_required ## Auditoria##
@permission_required("inventario.delete_clientes",raise_exception=True)
def eliminarclte(request):
    clte = None
    mensaje = ""
    try:
        clte = clientes.objects.get(id_compra = request.GET["idcliente"])
        y = maquinaria.objects.all()
        return render(request, "eliminarclientes.html", {"cltes" : clte, "y":y})
    except:
        clte = None

    if clte == None:
        id_compra = None
        try:
            id_compra = request.POST["id_compra"]
        except:
            id_compra = None

        if id_compra != None :
            clte = clientes.objects.get(id_compra = id_compra)

            rut_clt = request.POST["rut_cliente"]
            nombre_clt = request.POST["nombre_cliente"]
            telefono_clt = request.POST["telefono_contacto"]
            direccion_clt = request.POST["direccion"] 
            unidades_clt = request.POST["unidades"]
            fecha_clt = datetime.now()
            total_clt = request.POST ["total"]
            #maqui = request.GET["cod_maquina", False]
            #z = maquinaria.objects.get(activo__maquinaria__activo = maqui)
            #y = maquinaria.objects.all()            

            
            clte.rut_cliente = rut_clt
            clte.nombre_cliente = nombre_clt
            clte.telefono_contacto = telefono_clt
            clte.direccion = direccion_clt
            clte.unidades = unidades_clt
            clte.fecha = fecha_clt
            clte.total = total_clt
            #clte.maquinaria =z
            clte.activo= 'Inactivo'


            try:
                clte.save()
                mensaje = "Se ha Eliminado el cliente"
            except:
                mensaje = "Ha ocurrido un error al eliminar el cliente"
            
            return render(request, "eliminarclientes.html", {"mensaje":mensaje})
        else:
            mensaje = "No se ha encontrado al cliente consultado"

            return render(request, "eliminarclientes.html", {"mensaje":mensaje})
    else:
        mensaje = "No se ha encontrado al cliente consultado"

        return render(request, "eliminarclientes.html", {"mensaje":mensaje})
    
@login_required ##
@permission_required("inventario.activo",raise_exception=True)               
def actualizarusuario(request):
    return render(request,"actualizarusuario.html")

@login_required ##
@permission_required("inventario.change_usuario",raise_exception=True)
def editarus(request):
    us = None
    mensaje = ""

    try:
        us = Usuario.objects.get(username= request.GET["username"])
        return render(request, "actualizarusuario.html", {"us":us})

    except: 
        us = None
    
    if us == None:
        username = None
        try:
            username = request.POST["username"]
        except:
            username = None

        if username != None and username != "":
            us = Usuario.objects.get(username = username)

            nombres_a = request.POST["nombre_usuario"]
            apellidos_a = request.POST["primer_apellido"]
            rut_a = request.POST["rut_usuario"]
            email_a = request.POST["email_usuario"]
            rol_a = request.POST["rol"]

            us.first_name = nombres_a
            us.last_name= apellidos_a
            us.rut= rut_a
            us.email = email_a
            ##telefono_usuario = request.POST["telefono_usuario "]##
            us.rol = rol_a
            try:
                us.save()
                mensaje = "Se ha actualizado el usuario"
            except:
                mensaje = "Ha ocurrido un error al actualizar"


            return render(request, "actualizarusuario.html", {"mensaje":mensaje})
        
        else:
            mensaje = "No se encontro el usuario ingresado"

            return render(request, "actualizarusuario.html", {"mensaje":mensaje})
    else:
        mensaje = "No se encontró usuario"

        return render(request, "actualizarusuario.html", {"mensaje":mensaje})

@login_required ## 
def ver_auditoria(request):
    return render(request, "verauditoria.html")

@login_required ## Auditoria##
def ver_auditoria(request):
    # auditoria = LogEntry.objects.all()
    auditoria = LogEntry.objects.raw('SELECT * FROM auditlog_logentry')

    audDatos = []

    for aud in LogEntry.objects.raw('SELECT * FROM auditlog_logentry'):

        if 'last_login' in aud.changes_dict:
            aud.isLogin = True
        else:
            aud.isLogin = False

        if aud.actor_id != None:

            try:
                aud.user = Usuario.objects.get(id=int(aud.actor_id))
            except:
                aud.user = 'Usuario Eliminado'
        else:
            aud.user = Usuario.objects.get(id=int(aud.object_pk))
        # else:
        #     aud.user = 'Usuario no Registrado'

        log = json.loads(aud.changes)
        aud.log = log
        if 'is_active' in log:  
            if log['is_active'][0] and log['is_active'][0] != 'None':
                aud.action = 2
                aud.log = "Usuario Inactivo: " + aud.object_repr

        if 'activo' in log:  
            if log['activo'][0] and log['activo'][0] != 'None':
                aud.action = 2
                aud.log = "Inactivo: " + aud.object_pk   

        if aud.action == 0:
            aud.log = 'Creación de objeto: ' + aud.object_pk
        if aud.action == 1:
            aud.log = 'Atualización de objeto: ' + aud.object_pk
        if 'last_login' in log:
            aud.log = 'Inicio de sesión'
        audDatos.append(aud)

    datos = {
        'movimientos': audDatos
        # 'aud.changes_dict': aud.changes_dict
    }
    return render(request, "verauditoria.html", datos)